"""Scholar service — public API for loading, filtering, and updating scholar data.

Internal implementation is split across private sub-modules:
  _data.py         — unified scholars.json loading and annotation merging
  _filters.py      — multi-field filtering helpers
  _transformers.py — response shape converters (_to_list_item, _to_detail)
"""
from __future__ import annotations

import json
import logging
import math
import hashlib
import re
from collections import Counter
from datetime import UTC, datetime
from typing import Any

logger = logging.getLogger(__name__)

from app.services.core.institution.classification import normalize_org_type
from app.services.stores import scholar_annotation_store as annotation_store
from app.services.stores import supervised_student_store as student_store
from app.services.scholar._data import (
    SCHOLARS_FILE,
    _find_raw_file_by_hash,
    _merge_annotation,
    _load_all_with_annotations,
    _load_all_with_annotations_async,
)
from app.services.scholar._filters import (
    _apply_filters,
    get_institution_classification_map,
)
from app.services.scholar._transformers import _to_detail, _to_list_item
from app.services.scholar._create import import_scholars_excel, _parse_excel_row  # noqa: F401
from app.services.scholar._fast_query import query_scholar_list_fast


_MISSING_COLUMN_RE = re.compile(
    r'column\s+"(?P<column>[^"]+)"\s+of\s+relation\s+"scholars"\s+does\s+not\s+exist',
    re.IGNORECASE,
)


def _extract_missing_scholar_column(exc: Exception) -> str | None:
    match = _MISSING_COLUMN_RE.search(str(exc))
    if not match:
        return None
    return str(match.group("column") or "").strip() or None


async def _insert_scholar_with_schema_fallback(client: Any, payload: dict[str, Any]) -> None:
    """Insert scholar row with runtime schema fallback.

    Some environments have a lagging scholars table schema (missing optional
    columns like participated_event_ids/event_tags/project_tags). In that case
    we remove the missing column from insert payload and retry.
    """
    insert_payload = dict(payload)
    removed_columns: list[str] = []
    max_attempts = max(1, len(insert_payload))

    for _ in range(max_attempts):
        try:
            await client.table("scholars").insert(insert_payload).execute()
            if removed_columns:
                logger.warning(
                    "Inserted scholar after dropping missing columns: %s",
                    ",".join(removed_columns),
                )
            return
        except Exception as exc:
            missing_column = _extract_missing_scholar_column(exc)
            if not missing_column or missing_column not in insert_payload:
                raise
            insert_payload.pop(missing_column, None)
            removed_columns.append(missing_column)

    raise RuntimeError(
        "Unable to insert scholar due to unresolved schema mismatch after retries",
    )


def _normalize_project_tags(raw: Any) -> list[dict[str, str]]:
    tags: list[dict[str, str]] = []
    if not isinstance(raw, list):
        return tags
    for item in raw:
        if hasattr(item, "model_dump"):
            item = item.model_dump()
        if not isinstance(item, dict):
            continue
        category = str(item.get("category") or "").strip()
        subcategory = str(item.get("subcategory") or "").strip()
        if not category and not subcategory:
            continue
        tags.append(
            {
                "category": category,
                "subcategory": subcategory,
                "project_id": str(item.get("project_id") or ""),
                "project_title": str(item.get("project_title") or ""),
            }
        )
    return tags


def _normalize_event_tags(raw: Any) -> list[dict[str, str]]:
    tags: list[dict[str, str]] = []
    if not isinstance(raw, list):
        return tags
    for item in raw:
        if hasattr(item, "model_dump"):
            item = item.model_dump()
        if not isinstance(item, dict):
            continue
        category = str(item.get("category") or "").strip()
        series = str(item.get("series") or "").strip()
        event_type = str(item.get("event_type") or "").strip()
        if not category and not series and not event_type:
            continue
        tags.append(
            {
                "category": category,
                "series": series,
                "event_type": event_type,
                "event_id": str(item.get("event_id") or ""),
                "event_title": str(item.get("event_title") or ""),
            }
        )
    return tags


def _first_project_tag(tags: list[dict[str, str]]) -> tuple[str, str]:
    if not tags:
        return "", ""
    first = tags[0]
    return first.get("category", ""), first.get("subcategory", "")


def _derive_cobuild_from_tags(
    *,
    explicit: Any = None,
    project_tags: list[dict[str, str]] | None = None,
    event_tags: list[dict[str, str]] | None = None,
) -> bool:
    has_tags = bool(project_tags or []) or bool(event_tags or [])
    if has_tags:
        return True
    if isinstance(explicit, bool):
        return explicit
    return False


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _model_dump_maybe(value: Any) -> Any:
    if hasattr(value, "model_dump"):
        return value.model_dump()
    return value


def _split_people(raw: Any) -> list[str]:
    value = _model_dump_maybe(raw)
    if value is None:
        return []
    if isinstance(value, list):
        result: list[str] = []
        seen: set[str] = set()
        for item in value:
            token = _clean_text(item)
            if not token:
                continue
            key = token.lower()
            if key in seen:
                continue
            seen.add(key)
            result.append(token)
        return result
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return _split_people(parsed)
        except Exception:
            pass
        normalized = text
        for sep in ("；", ";", "，", ",", "、", "|", "/", "\n", "\r", "\t"):
            normalized = normalized.replace(sep, "|")
        tokens = [t.strip() for t in normalized.split("|") if t.strip()]
        return _split_people(tokens)
    return []


_YEAR_PATTERN = re.compile(r"(19|20)\d{2}")


def _to_year_int(raw: Any) -> int | None:
    text = _clean_text(raw)
    if not text:
        return None
    try:
        year = int(float(text))
    except Exception:
        match = _YEAR_PATTERN.search(text)
        if not match:
            return None
        year = int(match.group(0))
    if year < 1900 or year > 2100:
        return None
    return year


def _to_int(raw: Any, default: int = -1) -> int:
    try:
        return int(raw)
    except Exception:
        return default


def _stable_bigint(*parts: Any) -> int:
    joined = "|".join(_clean_text(p) for p in parts)
    digest = hashlib.sha256(joined.encode("utf-8")).digest()
    value = int.from_bytes(digest[:8], "big") & ((1 << 63) - 1)
    return value or 1


def _normalize_publication_item(raw: Any, scholar_id: str, idx: int) -> dict[str, Any] | None:
    item = _model_dump_maybe(raw)
    if not isinstance(item, dict):
        return None
    title = _clean_text(item.get("title"))
    if not title:
        return None
    venue = _clean_text(item.get("venue"))
    year = _to_year_int(item.get("year"))
    authors = _split_people(item.get("authors"))
    url = _clean_text(item.get("url"))
    citation_count = _to_int(item.get("citation_count"), default=-1)
    is_corresponding = bool(item.get("is_corresponding", False))
    added_by = _clean_text(item.get("added_by")) or "crawler"
    row_id = _stable_bigint(
        "pub",
        scholar_id,
        idx,
        title,
        venue,
        year or "",
        url,
        ",".join(authors),
    )
    return {
        "id": row_id,
        "scholar_id": scholar_id,
        "title": title,
        "venue": venue or None,
        "year": year,
        "authors": authors or None,
        "url": url or None,
        "citation_count": citation_count,
        "is_corresponding": is_corresponding,
        "added_by": added_by,
    }


def _normalize_patent_item(raw: Any, scholar_id: str, idx: int) -> dict[str, Any] | None:
    item = _model_dump_maybe(raw)
    if not isinstance(item, dict):
        return None
    title = _clean_text(item.get("title"))
    if not title:
        return None
    patent_no = _clean_text(item.get("patent_no"))
    year = _to_year_int(item.get("year"))
    inventors = _split_people(item.get("inventors"))
    patent_type = _clean_text(item.get("patent_type"))
    status = _clean_text(item.get("status"))
    added_by = _clean_text(item.get("added_by")) or "crawler"
    row_id = _stable_bigint(
        "pat",
        scholar_id,
        idx,
        title,
        patent_no,
        year or "",
        ",".join(inventors),
    )
    return {
        "id": row_id,
        "scholar_id": scholar_id,
        "title": title,
        "patent_no": patent_no or None,
        "year": year,
        "inventors": inventors or None,
        "patent_type": patent_type or None,
        "status": status or None,
        "added_by": added_by,
    }


def _publication_db_row_to_api(row: dict[str, Any]) -> dict[str, Any]:
    authors_raw = row.get("authors")
    if isinstance(authors_raw, list):
        authors = ", ".join(_clean_text(x) for x in authors_raw if _clean_text(x))
    else:
        authors = _clean_text(authors_raw)
    return {
        "title": _clean_text(row.get("title")),
        "venue": _clean_text(row.get("venue")),
        "year": "" if row.get("year") is None else str(row.get("year")),
        "authors": authors,
        "url": _clean_text(row.get("url")),
        "citation_count": _to_int(row.get("citation_count"), default=-1),
        "is_corresponding": bool(row.get("is_corresponding", False)),
        "added_by": _clean_text(row.get("added_by")) or "crawler",
    }


def _patent_db_row_to_api(row: dict[str, Any]) -> dict[str, Any]:
    inventors_raw = row.get("inventors")
    if isinstance(inventors_raw, list):
        inventors = ", ".join(_clean_text(x) for x in inventors_raw if _clean_text(x))
    else:
        inventors = _clean_text(inventors_raw)
    return {
        "title": _clean_text(row.get("title")),
        "patent_no": _clean_text(row.get("patent_no")),
        "year": "" if row.get("year") is None else str(row.get("year")),
        "inventors": inventors,
        "patent_type": _clean_text(row.get("patent_type")),
        "status": _clean_text(row.get("status")),
        "added_by": _clean_text(row.get("added_by")) or "crawler",
    }


async def _load_scholar_publications(scholar_id: str) -> list[dict[str, Any]]:
    try:
        from app.db.pool import get_pool  # noqa: PLC0415

        rows = await get_pool().fetch(
            """
            SELECT title, venue, year, authors, url, citation_count, is_corresponding, added_by
            FROM scholar_publications
            WHERE scholar_id = $1
            ORDER BY year DESC NULLS LAST, created_at DESC, id DESC
            """,
            scholar_id,
        )
        return [_publication_db_row_to_api(dict(r)) for r in rows]
    except Exception as exc:
        logger.warning("Failed to load scholar_publications for %s: %s", scholar_id, exc)
        return []


async def _load_scholar_patents(scholar_id: str) -> list[dict[str, Any]]:
    try:
        from app.db.pool import get_pool  # noqa: PLC0415

        rows = await get_pool().fetch(
            """
            SELECT title, patent_no, year, inventors, patent_type, status, added_by
            FROM scholar_patents
            WHERE scholar_id = $1
            ORDER BY year DESC NULLS LAST, created_at DESC, id DESC
            """,
            scholar_id,
        )
        return [_patent_db_row_to_api(dict(r)) for r in rows]
    except Exception as exc:
        logger.warning("Failed to load scholar_patents for %s: %s", scholar_id, exc)
        return []


async def _replace_scholar_publications(scholar_id: str, items: list[Any]) -> None:
    normalized_rows = [
        row
        for idx, item in enumerate(items)
        if (row := _normalize_publication_item(item, scholar_id, idx)) is not None
    ]
    from app.db.pool import get_pool  # noqa: PLC0415

    pool = get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.execute("DELETE FROM scholar_publications WHERE scholar_id = $1", scholar_id)
            if not normalized_rows:
                return
            await conn.executemany(
                """
                INSERT INTO scholar_publications
                (id, scholar_id, title, venue, year, authors, url, citation_count, is_corresponding, added_by)
                VALUES ($1, $2, $3, $4, $5, $6::text[], $7, $8, $9, $10)
                """,
                [
                    (
                        row["id"],
                        row["scholar_id"],
                        row["title"],
                        row["venue"],
                        row["year"],
                        row["authors"],
                        row["url"],
                        row["citation_count"],
                        row["is_corresponding"],
                        row["added_by"],
                    )
                    for row in normalized_rows
                ],
            )


async def _replace_scholar_patents(scholar_id: str, items: list[Any]) -> None:
    normalized_rows = [
        row
        for idx, item in enumerate(items)
        if (row := _normalize_patent_item(item, scholar_id, idx)) is not None
    ]
    from app.db.pool import get_pool  # noqa: PLC0415

    pool = get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.execute("DELETE FROM scholar_patents WHERE scholar_id = $1", scholar_id)
            if not normalized_rows:
                return
            await conn.executemany(
                """
                INSERT INTO scholar_patents
                (id, scholar_id, title, patent_no, year, inventors, patent_type, status, added_by)
                VALUES ($1, $2, $3, $4, $5, $6::text[], $7, $8, $9)
                """,
                [
                    (
                        row["id"],
                        row["scholar_id"],
                        row["title"],
                        row["patent_no"],
                        row["year"],
                        row["inventors"],
                        row["patent_type"],
                        row["status"],
                        row["added_by"],
                    )
                    for row in normalized_rows
                ],
            )


# ---------------------------------------------------------------------------
# Create operation (async, Supabase-first)
# ---------------------------------------------------------------------------


async def create_scholar(data: dict[str, Any]) -> tuple[dict[str, Any] | None, str]:
    """Create a new scholar record, saving to Supabase (with JSON fallback).

    Returns:
        (scholar_detail_dict, "")  on success
        (None, "duplicate:{url_hash}")  on duplicate
        (None, error_message)  on error
    """
    from app.crawlers.utils.dedup import compute_url_hash  # noqa: PLC0415

    name = (data.get("name") or "").strip()
    if not name:
        return None, "name is required"

    university = (data.get("university") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    profile_url = (data.get("profile_url") or "").strip()
    department = (data.get("department") or "").strip()

    dept_part = f"/{department}" if department else ""
    url = profile_url or f"manual://{name}@{university}{dept_part}"
    url_hash = compute_url_hash(url)

    # --- Try Supabase ---
    try:
        from app.db.client import get_client  # noqa: PLC0415
        client = get_client()

        # Duplicate check by url_hash (same URL = same scholar)
        dup_check = await client.table("scholars").select("id").eq("id", url_hash).execute()
        if dup_check.data:
            return None, f"duplicate:{url_hash}"

        # Also check same name + university combination
        name_check = await (
            client.table("scholars")
            .select("id,name,university,email,phone")
            .eq("name", name)
            .execute()
        )
        for existing in (name_check.data or []):
            if (existing.get("university") or "").lower() != university.lower():
                continue
            existing_email = (existing.get("email") or "").strip()
            existing_phone = (existing.get("phone") or "").strip()
            has_contact = bool(email or phone)
            existing_has_contact = bool(existing_email or existing_phone)
            if not has_contact and not existing_has_contact:
                return None, f"duplicate:{existing['id']}"
            if has_contact and existing_has_contact:
                if email and email.lower() == existing_email.lower():
                    return None, f"duplicate:{existing['id']}"
                if phone and phone == existing_phone:
                    return None, f"duplicate:{existing['id']}"

        project_tags = _normalize_project_tags(data.get("project_tags"))
        event_tags = _normalize_event_tags(data.get("event_tags"))
        first_project_category, first_project_subcategory = _first_project_tag(project_tags)
        participated_event_ids = data.get("participated_event_ids") or []

        record: dict[str, Any] = {
            "id": url_hash,
            "source_url": url,
            "name": name,
            "name_en": data.get("name_en") or "",
            "gender": data.get("gender") or "",
            "photo_url": data.get("photo_url") or "",
            "university": university,
            "department": department,
            "secondary_departments": data.get("secondary_departments") or [],
            "position": data.get("position") or "",
            "academic_titles": data.get("academic_titles") or [],
            "is_academician": bool(data.get("is_academician", False)),
            "research_areas": data.get("research_areas") or [],
            "keywords": data.get("keywords") or [],
            "bio": data.get("bio") or "",
            "bio_en": data.get("bio_en") or "",
            "email": email,
            "phone": phone,
            "office": data.get("office") or "",
            "profile_url": profile_url,
            "lab_url": data.get("lab_url") or "",
            "google_scholar_url": data.get("google_scholar_url") or "",
            "dblp_url": data.get("dblp_url") or "",
            "orcid": data.get("orcid") or "",
            "phd_institution": data.get("phd_institution") or "",
            "phd_year": int(data["phd_year"]) if data.get("phd_year") else None,
            "education": data.get("education") or [],
            "publications_count": -1,
            "h_index": -1,
            "citations_count": -1,
            "representative_publications": [],
            "patents": [],
            "awards": [],
            "is_advisor_committee": False,
            "adjunct_supervisor": {"status": "", "type": "", "agreement_type": "", "agreement_period": "", "recommender": ""},
            "is_potential_recruit": False,
            "institute_relation_notes": "",
            "supervised_students": [],
            "joint_research_projects": [],
            "joint_management_roles": [],
            "academic_exchange_records": [],
            "participated_event_ids": participated_event_ids,
            "event_tags": event_tags,
            "project_tags": project_tags,
            "is_cobuild_scholar": _derive_cobuild_from_tags(
                explicit=data.get("is_cobuild_scholar"),
                project_tags=project_tags,
                event_tags=event_tags,
            ),
            "relation_updated_by": "",
            "relation_updated_at": None,
            "recent_updates": [],
            "created_at": datetime.now(UTC).isoformat(),
            "updated_at": datetime.now(UTC).isoformat(),
            "metrics_updated_at": None,
            "project_category": first_project_category,
            "project_subcategory": first_project_subcategory,
        }
        insert_record = record
        try:
            from app.db.pool import get_pool  # noqa: PLC0415

            rows = await get_pool().fetch(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema='public' AND table_name='scholars'
                """,
            )
            scholar_cols = {str(r["column_name"]) for r in rows}
            if scholar_cols:
                insert_record = {k: v for k, v in record.items() if k in scholar_cols}
        except Exception as exc:
            logger.warning("Failed to introspect scholars columns, using full payload: %s", exc)

        await _insert_scholar_with_schema_fallback(client, insert_record)

        # Keep relation fields in annotation overlay as compatibility fallback
        # when runtime DB schema has not added all optional relation columns.
        relation_overlay: dict[str, Any] = {}
        if participated_event_ids:
            relation_overlay["participated_event_ids"] = participated_event_ids
        if event_tags:
            relation_overlay["event_tags"] = event_tags
        if project_tags:
            relation_overlay["project_tags"] = project_tags
        if project_tags or event_tags:
            relation_overlay["is_cobuild_scholar"] = _derive_cobuild_from_tags(
                explicit=data.get("is_cobuild_scholar"),
                project_tags=project_tags,
                event_tags=event_tags,
            )
        if relation_overlay:
            annotation_store.update_relation(url_hash, relation_overlay)

        return await get_scholar_detail(url_hash), ""

    except Exception as exc:
        logger.error("DB create_scholar failed: %s", exc, exc_info=True)
        return None, f"save failed: {exc}"


async def import_scholars_async(
    file_content: bytes,
    filename: str,
    skip_duplicates: bool = True,
    added_by: str = "user",
) -> dict[str, Any]:
    """Async version of Excel/CSV import — saves each record to Supabase via create_scholar.

    Returns a summary dict compatible with ScholarImportResult schema.
    """
    import csv  # noqa: PLC0415
    from io import StringIO, BytesIO  # noqa: PLC0415

    result: dict[str, Any] = {"total": 0, "success": 0, "skipped": 0, "failed": 0, "items": []}

    # --- Parse file ---
    try:
        is_csv = filename.lower().endswith(".csv")
        if is_csv:
            text = file_content.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            rows = list(reader)
        else:
            try:
                import openpyxl  # noqa: PLC0415
            except ImportError:
                result["failed"] = 1
                result["items"].append({"row": 0, "status": "failed", "name": "",
                                        "reason": "openpyxl not installed"})
                return result
            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(c).strip() if c else "" for c in header_row]
            rows = []
            for rv in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for ci, val in enumerate(rv):
                    if ci < len(headers):
                        row_dict[headers[ci]] = str(val).strip() if val else ""
                rows.append(row_dict)
    except Exception as exc:
        logger.error("Failed to parse file %s: %s", filename, exc, exc_info=True)
        result["failed"] = 1
        result["items"].append({"row": 0, "status": "failed", "name": "",
                                 "reason": f"parse error: {exc}"})
        return result

    result["total"] = len(rows)

    for row_idx, row in enumerate(rows, start=1):
        parsed = _parse_excel_row(row)
        name = parsed.get("name", "").strip()
        if not name:
            result["failed"] += 1
            result["items"].append({"row": row_idx, "status": "failed", "name": "",
                                     "reason": "name is required"})
            continue

        parsed["added_by"] = added_by
        detail, error = await create_scholar(parsed)

        if error.startswith("duplicate:"):
            existing_hash = error.split(":", 1)[1]
            if skip_duplicates:
                result["skipped"] += 1
                result["items"].append({"row": row_idx, "status": "skipped", "name": name,
                                         "url_hash": existing_hash, "reason": "duplicate"})
            else:
                result["failed"] += 1
                result["items"].append({"row": row_idx, "status": "failed", "name": name,
                                         "reason": "duplicate"})
        elif error:
            result["failed"] += 1
            result["items"].append({"row": row_idx, "status": "failed", "name": name,
                                     "reason": error})
        else:
            result["success"] += 1
            result["items"].append({"row": row_idx, "status": "success", "name": name,
                                     "url_hash": detail["url_hash"] if detail else ""})

    return result


async def batch_create_scholars(
    items: list[dict[str, Any]],
    skip_duplicates: bool = True,
    added_by: str = "user",
) -> dict[str, Any]:
    """Batch-create scholars from a list of dicts.  Each dict uses the same field
    names as ScholarCreateRequest.  Returns a summary dict.
    """
    result: dict[str, Any] = {"total": len(items), "success": 0, "skipped": 0, "failed": 0,
                               "items": []}
    for row_idx, data in enumerate(items, start=1):
        name = (data.get("name") or "").strip()
        if not name:
            result["failed"] += 1
            result["items"].append({"row": row_idx, "status": "failed", "name": "",
                                     "reason": "name is required"})
            continue

        data_with_by = {**data, "added_by": added_by}
        detail, error = await create_scholar(data_with_by)

        if error.startswith("duplicate:"):
            existing_hash = error.split(":", 1)[1]
            if skip_duplicates:
                result["skipped"] += 1
                result["items"].append({"row": row_idx, "status": "skipped", "name": name,
                                         "url_hash": existing_hash, "reason": "duplicate"})
            else:
                result["failed"] += 1
                result["items"].append({"row": row_idx, "status": "failed", "name": name,
                                         "reason": "duplicate"})
        elif error:
            result["failed"] += 1
            result["items"].append({"row": row_idx, "status": "failed", "name": name,
                                     "reason": error})
        else:
            result["success"] += 1
            result["items"].append({"row": row_idx, "status": "success", "name": name,
                                     "url_hash": detail["url_hash"] if detail else ""})

    return result


# ---------------------------------------------------------------------------
# Read operations
# ---------------------------------------------------------------------------


async def _resolve_institution_names_by_group_or_category(
    group: str | None,
    category: str | None,
) -> list[str] | None:
    """通过 institution_group 或 institution_category 查找机构名称列表，用于学者过滤."""
    if not group and not category:
        return None
    try:
        from app.services.core.institution import get_institution_list  # noqa: PLC0415

        page = 1
        page_size = 1000
        names: list[str] = []
        seen: set[str] = set()
        while True:
            result = await get_institution_list(
                group=group,
                category=category,
                entity_type="organization",
                page=page,
                page_size=page_size,
            )
            for item in result.items:
                name = str(getattr(item, "name", "") or "").strip()
                if not name or name in seen:
                    continue
                seen.add(name)
                names.append(name)
            if page >= max(int(result.total_pages or 1), 1):
                break
            page += 1

        if not names:
            # Be tolerant to frontend enum mismatches (e.g. "全部"/unknown values):
            # do not turn the whole scholar list into 0 by applying an empty name-set.
            logger.warning(
                "No institutions resolved for group=%s category=%s, skip institution_names filter",
                group,
                category,
            )
            return None
        return names
    except Exception as exc:
        logger.warning("Failed to resolve institution names for group=%s category=%s: %s", group, category, exc)
        return None


def _normalize_optional_filter_text(value: str | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text in {"全部", "all", "ALL", "All", "不限"}:
        return None
    return text


async def get_scholar_list(
    *,
    university: str | None = None,
    department: str | None = None,
    position: str | None = None,
    is_academician: bool | None = None,
    is_potential_recruit: bool | None = None,
    is_advisor_committee: bool | None = None,
    is_adjunct_supervisor: bool | None = None,
    has_email: bool | None = None,
    keyword: str | None = None,
    community_name: str | None = None,
    community_type: str | None = None,
    project_category: str | None = None,
    project_subcategory: str | None = None,
    project_categories: str | None = None,
    project_subcategories: str | None = None,
    event_types: str | None = None,
    participated_event_id: str | None = None,
    is_cobuild_scholar: bool | None = None,
    region: str | None = None,
    affiliation_type: str | None = None,
    institution_group: str | None = None,
    institution_category: str | None = None,
    page: int = 1,
    page_size: int = 20,
    custom_field_key: str | None = None,
    custom_field_value: str | None = None,
) -> dict[str, Any]:
    university = _normalize_optional_filter_text(university)
    department = _normalize_optional_filter_text(department)
    position = _normalize_optional_filter_text(position)
    keyword = _normalize_optional_filter_text(keyword)
    institution_group = _normalize_optional_filter_text(institution_group)
    institution_category = _normalize_optional_filter_text(institution_category)

    affiliation_type = normalize_org_type(affiliation_type)

    # Resolve institution_group/category → list of institution names for filtering
    institution_names: list[str] | None = None
    if institution_group or institution_category:
        institution_names = await _resolve_institution_names_by_group_or_category(
            institution_group, institution_category
        )

    # Primary path: SQL pushdown + pagination (much faster than loading all rows).
    # Multi-value tag filters are currently implemented in in-memory fallback path.
    has_multi_tag_filters = bool(project_categories or project_subcategories or event_types)
    has_region_or_type_filters = bool(region or affiliation_type)
    # University/department filters use affiliation normalization logic in
    # fallback path to stay consistent with institution-tree aggregation.
    has_affiliation_text_filters = bool(university or department)
    if not has_multi_tag_filters and not has_region_or_type_filters and not has_affiliation_text_filters:
        try:
            return await query_scholar_list_fast(
                university=university,
                department=department,
                position=position,
                is_academician=is_academician,
                is_potential_recruit=is_potential_recruit,
                is_advisor_committee=is_advisor_committee,
                is_adjunct_supervisor=is_adjunct_supervisor,
                has_email=has_email,
                keyword=keyword,
                community_name=community_name,
                community_type=community_type,
                project_category=project_category,
                project_subcategory=project_subcategory,
                participated_event_id=participated_event_id,
                is_cobuild_scholar=is_cobuild_scholar,
                region=region,
                affiliation_type=affiliation_type,
                institution_names=institution_names,
                custom_field_key=custom_field_key,
                custom_field_value=custom_field_value,
                page=page,
                page_size=page_size,
            )
        except Exception as exc:
            logger.warning("Fast scholar list query failed, fallback to legacy path: %s", exc)

    # Fallback path: in-memory filtering over full dataset.
    items = await _load_all_with_annotations_async()

    inst_map: dict = {}
    if region or affiliation_type:
        inst_map = await get_institution_classification_map()

    filtered = _apply_filters(
        items,
        university=university,
        department=department,
        position=position,
        is_academician=is_academician,
        is_potential_recruit=is_potential_recruit,
        is_advisor_committee=is_advisor_committee,
        is_adjunct_supervisor=is_adjunct_supervisor,
        has_email=has_email,
        keyword=keyword,
        community_name=community_name,
        community_type=community_type,
        project_category=project_category,
        project_subcategory=project_subcategory,
        project_categories=project_categories,
        project_subcategories=project_subcategories,
        event_types=event_types,
        participated_event_id=participated_event_id,
        is_cobuild_scholar=is_cobuild_scholar,
        region=region,
        affiliation_type=affiliation_type,
        institution_names=institution_names,
        custom_field_key=custom_field_key,
        custom_field_value=custom_field_value,
        inst_map=inst_map,
    )

    filtered.sort(key=lambda i: i.get("name", ""))

    total = len(filtered)
    total_pages = math.ceil(total / page_size) if total > 0 else 1
    # Be tolerant to stale frontend page index (e.g. keeping page=14 after filters change).
    effective_page = min(max(page, 1), total_pages)
    start = (effective_page - 1) * page_size
    page_items = filtered[start : start + page_size]

    return {
        "total": total,
        "page": effective_page,
        "page_size": page_size,
        "total_pages": total_pages,
        "items": [_to_list_item(i) for i in page_items],
    }


async def get_scholar_detail(url_hash: str) -> dict[str, Any] | None:
    """Return full scholar detail merged with annotations, or None if not found."""
    # Primary path: direct DB point query by primary key.
    try:
        from app.db.client import get_client  # noqa: PLC0415

        client = get_client()
        resp = await client.table("scholars").select("*").eq("id", url_hash).limit(1).execute()
        if resp.data:
            item = dict(resp.data[0])
            if "url_hash" not in item:
                item["url_hash"] = item.get("id", "")
            if "url" not in item:
                item["url"] = item.get("source_url", "")
            ann = annotation_store.get_annotation(url_hash)
            if ann:
                _merge_annotation(item, ann)
            detail = _to_detail(item)
            detail["supervised_students_count"] = await student_store.count_students(url_hash)
            publications = await _load_scholar_publications(url_hash)
            patents = await _load_scholar_patents(url_hash)
            if publications:
                detail["representative_publications"] = publications
            if patents:
                detail["patents"] = patents
            return detail
    except Exception as exc:
        logger.warning("Fast scholar detail query failed, fallback to legacy path: %s", exc)

    # Fallback path: legacy full-data scan.
    items = await _load_all_with_annotations_async()
    for item in items:
        if item.get("url_hash", "") == url_hash:
            detail = _to_detail(item)
            detail["supervised_students_count"] = await student_store.count_students(url_hash)
            publications = await _load_scholar_publications(url_hash)
            patents = await _load_scholar_patents(url_hash)
            if publications:
                detail["representative_publications"] = publications
            if patents:
                detail["patents"] = patents
            return detail
    return None


async def get_scholar_stats(
    *,
    university: str | None = None,
    department: str | None = None,
    position: str | None = None,
    is_academician: bool | None = None,
    is_potential_recruit: bool | None = None,
    is_advisor_committee: bool | None = None,
    is_adjunct_supervisor: bool | None = None,
    has_email: bool | None = None,
    keyword: str | None = None,
    community_name: str | None = None,
    community_type: str | None = None,
    project_category: str | None = None,
    project_subcategory: str | None = None,
    project_categories: str | None = None,
    project_subcategories: str | None = None,
    event_types: str | None = None,
    participated_event_id: str | None = None,
    is_cobuild_scholar: bool | None = None,
    region: str | None = None,
    affiliation_type: str | None = None,
    institution_group: str | None = None,
    institution_category: str | None = None,
    custom_field_key: str | None = None,
    custom_field_value: str | None = None,
) -> dict[str, Any]:
    """Get scholar statistics with optional filtering.

    Applies the same filters as get_scholar_list() to ensure consistency
    between list view and stats view.
    """
    university = _normalize_optional_filter_text(university)
    department = _normalize_optional_filter_text(department)
    position = _normalize_optional_filter_text(position)
    keyword = _normalize_optional_filter_text(keyword)
    institution_group = _normalize_optional_filter_text(institution_group)
    institution_category = _normalize_optional_filter_text(institution_category)

    affiliation_type = normalize_org_type(affiliation_type)

    # Resolve institution_group/category → list of institution names for filtering
    institution_names: list[str] | None = None
    if institution_group or institution_category:
        institution_names = await _resolve_institution_names_by_group_or_category(
            institution_group, institution_category
        )

    items = await _load_all_with_annotations_async()

    # Fetch DB-based classification map when region/affiliation_type filter is active
    inst_map: dict = {}
    if region or affiliation_type:
        inst_map = await get_institution_classification_map()

    # Apply the same filters as get_scholar_list()
    filtered = _apply_filters(
        items,
        university=university,
        department=department,
        position=position,
        is_academician=is_academician,
        is_potential_recruit=is_potential_recruit,
        is_advisor_committee=is_advisor_committee,
        is_adjunct_supervisor=is_adjunct_supervisor,
        has_email=has_email,
        keyword=keyword,
        community_name=community_name,
        community_type=community_type,
        project_category=project_category,
        project_subcategory=project_subcategory,
        project_categories=project_categories,
        project_subcategories=project_subcategories,
        event_types=event_types,
        participated_event_id=participated_event_id,
        is_cobuild_scholar=is_cobuild_scholar,
        region=region,
        affiliation_type=affiliation_type,
        institution_names=institution_names,
        custom_field_key=custom_field_key,
        custom_field_value=custom_field_value,
        inst_map=inst_map,
    )

    # Calculate stats based on filtered items
    total = len(filtered)
    academicians = sum(1 for i in filtered if i.get("is_academician", False))
    potential_recruits = sum(1 for i in filtered if i.get("is_potential_recruit", False))
    advisor_committee = sum(1 for i in filtered if i.get("is_advisor_committee", False))
    adjunct_supervisors = sum(
        1 for i in filtered
        if (i.get("adjunct_supervisor") or {}).get("status")
    )

    uni_counter: Counter = Counter()
    dept_counter: Counter = Counter()
    pos_counter: Counter = Counter()

    for item in filtered:
        uni = item.get("university", "") or "未知"
        uni_counter[uni] += 1
        dept = item.get("department", "") or "未知"
        dept_counter[(uni, dept)] += 1
        pos = item.get("position", "") or "未知"
        pos_counter[pos] += 1

    return {
        "total": total,
        "academicians": academicians,
        "potential_recruits": potential_recruits,
        "advisor_committee": advisor_committee,
        "adjunct_supervisors": adjunct_supervisors,
        "by_university": [
            {"university": u, "count": c}
            for u, c in uni_counter.most_common()  # 返回所有机构，不限制数量
        ],
        "by_department": [
            {"university": u, "department": d, "count": c}
            for (u, d), c in dept_counter.most_common()  # 返回所有院系
        ],
        "by_position": [
            {"position": p, "count": c}
            for p, c in pos_counter.most_common()  # 返回所有职称
        ],
    }


# ---------------------------------------------------------------------------
# Write helpers (delegate to annotation_store)
# ---------------------------------------------------------------------------


async def update_scholar_relation(url_hash: str, updates: dict[str, Any]) -> dict[str, Any] | None:
    if await get_scholar_detail(url_hash) is None:
        return None
    normalized_updates: dict[str, Any] = {}
    for key, value in updates.items():
        if hasattr(value, "model_dump"):
            normalized_updates[key] = value.model_dump()
            continue
        if isinstance(value, list):
            normalized_updates[key] = [
                item.model_dump() if hasattr(item, "model_dump") else item
                for item in value
            ]
            continue
        normalized_updates[key] = value

    if "project_tags" in normalized_updates:
        tags = _normalize_project_tags(normalized_updates.get("project_tags"))
        normalized_updates["project_tags"] = tags
        first_category, first_subcategory = _first_project_tag(tags)
        normalized_updates["project_category"] = first_category
        normalized_updates["project_subcategory"] = first_subcategory

    if "event_tags" in normalized_updates:
        normalized_updates["event_tags"] = _normalize_event_tags(
            normalized_updates.get("event_tags")
        )

    if (
        "project_tags" in normalized_updates
        or "event_tags" in normalized_updates
    ):
        normalized_updates["is_cobuild_scholar"] = _derive_cobuild_from_tags(
            explicit=normalized_updates.get("is_cobuild_scholar"),
            project_tags=normalized_updates.get("project_tags"),
            event_tags=normalized_updates.get("event_tags"),
        )

    normalized_updates["relation_updated_at"] = datetime.now(UTC).isoformat()

    # Persist to scholars table when possible (dedicated columns).
    try:
        from app.db.client import get_client  # noqa: PLC0415

        client = get_client()
        await client.table("scholars").update(normalized_updates).eq("id", url_hash).execute()
    except Exception as exc:
        logger.warning("DB update_scholar_relation failed, fallback to annotations only: %s", exc)

    # Keep annotation overlay for backward compatibility.
    annotation_store.update_relation(url_hash, normalized_updates)
    return await get_scholar_detail(url_hash)


async def add_scholar_update(url_hash: str, update: dict[str, Any]) -> dict[str, Any] | None:
    if await get_scholar_detail(url_hash) is None:
        return None
    annotation_store.add_user_update(url_hash, update)
    return await get_scholar_detail(url_hash)


async def delete_scholar_update(url_hash: str, update_idx: int) -> dict[str, Any] | None:
    if await get_scholar_detail(url_hash) is None:
        return None
    annotation_store.delete_user_update(url_hash, update_idx)
    return await get_scholar_detail(url_hash)


async def update_scholar_achievements(url_hash: str, updates: dict[str, Any]) -> dict[str, Any] | None:
    if await get_scholar_detail(url_hash) is None:
        return None
    normalized_updates: dict[str, Any] = {}
    for key, value in updates.items():
        if hasattr(value, "model_dump"):
            normalized_updates[key] = value.model_dump()
        elif isinstance(value, list):
            normalized_updates[key] = [
                item.model_dump() if hasattr(item, "model_dump") else item
                for item in value
            ]
        else:
            normalized_updates[key] = value

    # Move representative publications / patents into dedicated relation tables.
    if "representative_publications" in normalized_updates:
        try:
            await _replace_scholar_publications(
                url_hash,
                normalized_updates.get("representative_publications") or [],
            )
        except Exception as exc:
            logger.warning("Failed updating scholar_publications for %s: %s", url_hash, exc)
    if "patents" in normalized_updates:
        try:
            await _replace_scholar_patents(
                url_hash,
                normalized_updates.get("patents") or [],
            )
        except Exception as exc:
            logger.warning("Failed updating scholar_patents for %s: %s", url_hash, exc)

    # Keep legacy columns for compatibility.
    db_patch: dict[str, Any] = {"updated_at": datetime.now(UTC).isoformat()}
    for key in (
        "representative_publications",
        "patents",
        "awards",
        "h_index",
        "citations_count",
        "publications_count",
    ):
        if key in normalized_updates:
            db_patch[key] = normalized_updates[key]

    try:
        from app.db.client import get_client  # noqa: PLC0415

        client = get_client()
        await client.table("scholars").update(db_patch).eq("id", url_hash).execute()
    except Exception as exc:
        logger.warning("DB update_scholar_achievements failed for %s: %s", url_hash, exc)

    annotation_store.update_achievements(url_hash, normalized_updates)
    return await get_scholar_detail(url_hash)


# ---------------------------------------------------------------------------
# Raw JSON modification helpers
# ---------------------------------------------------------------------------


async def update_scholar_basic(url_hash: str, updates: dict[str, Any]) -> dict[str, Any] | None:
    """Update basic scholar information.

    Tries Supabase DB first (mirrors the read path), falls back to local JSON file.
    Returns the updated full detail (merged with annotations) or None if not found.
    """
    # Build the payload to write (exclude meta-param, add audit fields)
    db_updates: dict[str, Any] = {}
    for key, value in updates.items():
        if key == "updated_by":
            continue
        if isinstance(value, list):
            db_updates[key] = [
                v.model_dump() if hasattr(v, "model_dump") else v
                for v in value
            ]
        else:
            db_updates[key] = value
    db_updates["updated_at"] = datetime.now(UTC).isoformat()

    # custom_fields 浅合并
    if "custom_fields" in db_updates:
        from app.services.core.custom_fields import apply_custom_fields_update  # noqa: PLC0415
        try:
            from app.db.client import get_client as _gc  # noqa: PLC0415
            cur = await _gc().table("scholars").select("custom_fields").eq("id", url_hash).execute()
            if cur.data:
                apply_custom_fields_update(db_updates, cur.data[0])
        except Exception:
            pass  # fallback: write as-is

    # --- Try Supabase first (matches the read path) ---
    try:
        from app.db.client import get_client  # noqa: PLC0415
        client = get_client()
        # Check existence first (supabase-py v2 update() returns [] by default without .select())
        exist = await client.table("scholars").select("id").eq("id", url_hash).execute()
        if exist.data:
            await client.table("scholars").update(db_updates).eq("id", url_hash).execute()
            return await get_scholar_detail(url_hash)
        # Not in DB → fall through to JSON fallback
    except Exception as exc:
        logger.warning("Supabase update_scholar_basic failed, trying local JSON: %s", exc)

    # --- Fallback: local JSON file ---
    result = _find_raw_file_by_hash(url_hash)
    if result is None:
        return None

    file_path, item_idx = result

    with open(file_path, encoding="utf-8") as f:
        data = json.load(f)

    scholar = data["scholars"][item_idx]
    scholar.update(db_updates)
    data["scholars"][item_idx] = scholar

    tmp_path = file_path.with_suffix(".tmp")
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    tmp_path.replace(file_path)

    return await get_scholar_detail(url_hash)


async def delete_scholar(url_hash: str) -> bool:
    """Delete a scholar record (Supabase-first, JSON fallback)."""
    # --- Try Supabase first ---
    try:
        from app.db.client import get_client  # noqa: PLC0415
        client = get_client()
        # supabase-py v2 delete() also returns [] by default without .select()
        exist = await client.table("scholars").select("id").eq("id", url_hash).execute()
        if exist.data:
            await client.table("scholars").delete().eq("id", url_hash).execute()
            annotation_store.delete_all_for_faculty(url_hash)
            await student_store.delete_all_students(url_hash)
            return True
        # Not in DB → fall through to JSON
    except Exception as exc:
        logger.warning("DB delete_scholar failed, trying JSON: %s", exc)

    # --- Fallback: local JSON file ---
    try:
        result = _find_raw_file_by_hash(url_hash)
        if result is None:
            return False

        file_path, item_idx = result

        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)

        scholars = data.get("scholars", [])
        if item_idx >= len(scholars) or scholars[item_idx].get("url_hash") != url_hash:
            return False

        scholars.pop(item_idx)
        data["scholars"] = scholars
        data["last_updated"] = datetime.now(UTC).isoformat()

        tmp_path = file_path.with_suffix(".tmp")
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        tmp_path.replace(file_path)

        annotation_store.delete_all_for_faculty(url_hash)
        await student_store.delete_all_students(url_hash)
        return True

    except Exception as exc:
        logger.error("Error deleting scholar %s: %s", url_hash, exc, exc_info=True)
        return False
