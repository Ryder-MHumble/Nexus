"""Scholar creation service — manual scholar creation and Excel bulk import."""
from __future__ import annotations

import csv
import json
import logging
from datetime import UTC, datetime
from io import StringIO
from pathlib import Path
from typing import Any

from app.crawlers.utils.dedup import compute_url_hash
from app.services.scholar._data import SCHOLARS_FILE, _load_all_with_annotations

logger = logging.getLogger(__name__)

_EMPTY_ADJUNCT: dict[str, str] = {
    "status": "", "type": "", "agreement_type": "",
    "agreement_period": "", "recommender": "",
}


def _generate_url_for_scholar(name: str, university: str, department: str, profile_url: str) -> str:
    if profile_url:
        return profile_url
    dept_part = f"/{department}" if department else ""
    return f"manual://{name}@{university}{dept_part}"


def _check_duplicate(name: str, university: str, email: str, phone: str) -> tuple[bool, str]:
    all_scholars = _load_all_with_annotations()
    for scholar in all_scholars:
        existing_name = scholar.get("name", "").strip()
        existing_uni = scholar.get("university", "").strip()
        existing_email = scholar.get("email", "").strip()
        existing_phone = scholar.get("phone", "").strip()
        if existing_name.lower() != name.lower():
            continue
        if existing_uni.lower() != university.lower():
            continue
        has_contact = bool(email or phone)
        existing_has_contact = bool(existing_email or existing_phone)
        if not has_contact and not existing_has_contact:
            return True, scholar.get("url_hash", "")
        if has_contact and existing_has_contact:
            if email and email.lower() == existing_email.lower():
                return True, scholar.get("url_hash", "")
            if phone and phone == existing_phone:
                return True, scholar.get("url_hash", "")
    return False, ""


def _save_scholar(record: dict[str, Any]) -> None:
    """Append a new scholar record to scholars.json (atomic write)."""
    if SCHOLARS_FILE.exists():
        with open(SCHOLARS_FILE, encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {"last_updated": "", "scholars": []}

    data["scholars"].append(record)
    data["last_updated"] = datetime.now(UTC).isoformat()

    tmp = SCHOLARS_FILE.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    tmp.replace(SCHOLARS_FILE)


def create_scholar(data: dict[str, Any]) -> tuple[dict[str, Any] | None, str]:
    """Create a new scholar record manually.

    Returns:
        (scholar_detail_dict, error_message)
        - On success: (detail, "")
        - On duplicate: (None, "duplicate:{url_hash}")
        - On error: (None, error_message)
    """
    name = data.get("name", "").strip()
    university = data.get("university", "").strip()
    email = data.get("email", "").strip()
    phone = data.get("phone", "").strip()
    profile_url = data.get("profile_url", "").strip()

    if not name:
        return None, "name is required"

    is_dup, existing_hash = _check_duplicate(name, university, email, phone)
    if is_dup:
        return None, f"duplicate:{existing_hash}"

    url = _generate_url_for_scholar(name, university, data.get("department", ""), profile_url)
    url_hash = compute_url_hash(url)
    project_tags_raw = data.get("project_tags") or []
    project_tags: list[dict[str, str]] = []
    if isinstance(project_tags_raw, list):
        for tag in project_tags_raw:
            if not isinstance(tag, dict):
                continue
            category = str(tag.get("category") or "").strip()
            subcategory = str(tag.get("subcategory") or "").strip()
            if not category and not subcategory:
                continue
            project_tags.append(
                {
                    "category": category,
                    "subcategory": subcategory,
                    "project_id": str(tag.get("project_id") or ""),
                    "project_title": str(tag.get("project_title") or ""),
                }
            )
    first_project_category = project_tags[0]["category"] if project_tags else ""
    first_project_subcategory = project_tags[0]["subcategory"] if project_tags else ""

    record: dict[str, Any] = {
        "url_hash": url_hash,
        "url": url,
        "content": "",
        "tags": [],
        # Basic identity
        "name": name,
        "name_en": data.get("name_en", ""),
        "gender": data.get("gender", ""),
        "photo_url": data.get("photo_url", ""),
        # Affiliation
        "university": university,
        "department": data.get("department", ""),
        "secondary_departments": data.get("secondary_departments") or [],
        # Academic profile
        "position": data.get("position", ""),
        "academic_titles": data.get("academic_titles") or [],
        "is_academician": bool(data.get("is_academician", False)),
        "research_areas": data.get("research_areas") or [],
        "keywords": data.get("keywords") or [],
        "bio": data.get("bio", ""),
        "bio_en": data.get("bio_en", ""),
        # Contact
        "email": email,
        "phone": phone,
        "office": data.get("office", ""),
        "profile_url": profile_url,
        "lab_url": data.get("lab_url", ""),
        "google_scholar_url": data.get("google_scholar_url", ""),
        "dblp_url": data.get("dblp_url", ""),
        "orcid": data.get("orcid", ""),
        # Education
        "phd_institution": data.get("phd_institution", ""),
        "phd_year": data.get("phd_year", ""),
        "education": data.get("education") or [],
        # Metrics (support enriched data)
        "publications_count": data.get("publications_count", -1),
        "h_index": data.get("h_index", -1),
        "citations_count": data.get("citations_count", -1),
        "metrics_updated_at": "",
        # Achievements (support enriched data)
        "representative_publications": data.get("representative_publications") or [],
        "patents": data.get("patents") or [],
        "awards": data.get("awards") or [],
        # Institute relations
        "is_advisor_committee": False,
        "adjunct_supervisor": dict(_EMPTY_ADJUNCT),
        "is_potential_recruit": False,
        "institute_relation_notes": "",
        "supervised_students": [],
        "joint_research_projects": [],
        "joint_management_roles": [],
        "academic_exchange_records": [],
        "participated_event_ids": data.get("participated_event_ids") or [],
        "event_tags": data.get("event_tags") or [],
        "project_tags": project_tags,
        "is_cobuild_scholar": bool(data.get("is_cobuild_scholar", bool(project_tags))),
        "relation_updated_by": "",
        "relation_updated_at": "",
        "recent_updates": [],
        # Custom fields (support enriched data)
        "custom_fields": data.get("custom_fields") or {},
        "project_category": first_project_category,
        "project_subcategory": first_project_subcategory,
    }

    try:
        _save_scholar(record)
    except Exception as exc:
        logger.error("Failed to save scholar %s: %s", name, exc, exc_info=True)
        return None, f"save failed: {exc}"

    from app.services.scholar import get_scholar_detail
    return get_scholar_detail(url_hash), ""


# ---------------------------------------------------------------------------
# Excel import (unchanged logic, reuses create_scholar above)
# ---------------------------------------------------------------------------

_COLUMN_MAP = {
    "姓名": "name", "name": "name",
    "英文名": "name_en", "name_en": "name_en",
    "性别": "gender", "gender": "gender",
    "照片": "photo_url", "photo": "photo_url", "photo_url": "photo_url",
    "所属院校": "university", "所属高校": "university", "所属机构": "university",
    "院校": "university", "高校": "university", "大学": "university", "university": "university",
    "院系/部门": "department", "所属院系": "department", "部门": "department",
    "院系": "department", "department": "department",
    "兼职院系": "secondary_departments", "secondary_departments": "secondary_departments",
    "职称": "position", "position": "position",
    "学术头衔": "academic_titles", "academic_titles": "academic_titles",
    "是否院士": "is_academician", "is_academician": "is_academician",
    "研究方向": "research_areas", "research_areas": "research_areas",
    "关键词": "keywords", "keywords": "keywords",
    "简介": "bio", "bio": "bio",
    "英文简介": "bio_en", "bio_en": "bio_en",
    "邮箱": "email", "email": "email",
    "电话": "phone", "phone": "phone",
    "办公室": "office", "office": "office",
    "主页": "profile_url", "个人主页": "profile_url", "profile_url": "profile_url",
    "实验室主页": "lab_url", "lab_url": "lab_url",
    "谷歌学术": "google_scholar_url", "google scholar": "google_scholar_url", "google_scholar_url": "google_scholar_url",
    "dblp": "dblp_url", "dblp_url": "dblp_url",
    "orcid": "orcid",
    "博士院校": "phd_institution", "phd_institution": "phd_institution",
    "博士年份": "phd_year", "phd_year": "phd_year",
    "教育经历": "education", "education": "education",
    "h指数": "h_index", "h_index": "h_index",
    "被引次数": "citations_count", "citations_count": "citations_count",
    "论文数": "publications_count", "publications_count": "publications_count",
    "代表性论文": "representative_publications", "representative_publications": "representative_publications",
    "专利": "patents", "patents": "patents",
    "获奖": "awards", "awards": "awards",
    "自定义字段": "custom_fields", "custom_fields": "custom_fields",
}


def _normalize_column_name(col: str) -> str:
    return _COLUMN_MAP.get(col.strip().lower(), "")


def _parse_bool(value: str) -> bool:
    return value.strip().lower() in ("是", "true", "1", "yes")


def _parse_list(value: str, delimiter: str = ",") -> list[str]:
    if not value:
        return []
    if ";" in value:
        delimiter = ";"
    return [item.strip() for item in value.split(delimiter) if item.strip()]


def _parse_json_field(value: str) -> Any:
    """Parse JSON string field (for education, awards, publications, custom_fields)."""
    if not value or not value.strip():
        return None
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        logger.warning(f"Failed to parse JSON field: {value[:100]}")
        return None


def _parse_int(value: str) -> int:
    """Parse integer field with fallback to -1."""
    if not value or not value.strip():
        return -1
    try:
        return int(value)
    except ValueError:
        return -1


def _parse_excel_row(row: dict[str, str]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for col, value in row.items():
        field = _normalize_column_name(col)
        if not field:
            continue
        if not value or (isinstance(value, str) and not value.strip()):
            continue

        value = value.strip() if isinstance(value, str) else value

        # Boolean fields
        if field == "is_academician":
            result[field] = _parse_bool(value)
        # List fields
        elif field in ("academic_titles", "research_areas", "keywords", "secondary_departments"):
            result[field] = _parse_list(value)
        # Integer fields
        elif field in ("h_index", "citations_count", "publications_count"):
            result[field] = _parse_int(value)
        # JSON fields
        elif field in ("education", "representative_publications", "patents", "awards", "custom_fields"):
            parsed = _parse_json_field(value)
            if parsed is not None:
                result[field] = parsed
        # String fields
        else:
            result[field] = value
    return result


def import_scholars_excel(
    file_content: bytes,
    filename: str,
    added_by: str = "user",
    skip_duplicates: bool = True,
) -> dict[str, Any]:
    result = {"total": 0, "success": 0, "skipped": 0, "failed": 0, "items": []}
    is_csv = filename.lower().endswith(".csv")

    try:
        if is_csv:
            text = file_content.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            rows = list(reader)
        else:
            try:
                import openpyxl
            except ImportError:
                result["failed"] = 1
                result["items"].append({
                    "row": 0, "status": "failed", "name": "",
                    "reason": "openpyxl not installed (run: pip install openpyxl)",
                })
                return result

            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell).strip() if cell else "" for cell in header_row]
            rows = []
            for row_values in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for col_idx, value in enumerate(row_values):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = str(value).strip() if value else ""
                rows.append(row_dict)
    except Exception as exc:
        logger.error("Failed to parse file %s: %s", filename, exc, exc_info=True)
        result["failed"] = 1
        result["items"].append({"row": 0, "status": "failed", "name": "", "reason": f"parse error: {exc}"})
        return result

    result["total"] = len(rows)

    for row_idx, row in enumerate(rows, start=1):
        parsed = _parse_excel_row(row)
        name = parsed.get("name", "").strip()
        if not name:
            result["failed"] += 1
            result["items"].append({"row": row_idx, "status": "failed", "name": "", "reason": "name is required"})
            continue

        parsed["added_by"] = added_by
        detail, error = create_scholar(parsed)

        if error.startswith("duplicate:"):
            existing_hash = error.split(":", 1)[1]
            if skip_duplicates:
                result["skipped"] += 1
                result["items"].append({
                    "row": row_idx, "status": "skipped", "name": name,
                    "url_hash": existing_hash, "reason": "duplicate",
                })
            else:
                result["failed"] += 1
                result["items"].append({"row": row_idx, "status": "failed", "name": name, "reason": "duplicate"})
        elif error:
            result["failed"] += 1
            result["items"].append({"row": row_idx, "status": "failed", "name": name, "reason": error})
        else:
            result["success"] += 1
            result["items"].append({
                "row": row_idx, "status": "success", "name": name,
                "url_hash": detail["url_hash"] if detail else "",
            })

    return result
